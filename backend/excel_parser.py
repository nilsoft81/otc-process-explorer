"""Validate and parse Process Design Excel into ALL_PROCESSES list.

Layout (Detailed Process Flows format):
  Row 5 (idx 4): main headers
  Row 6 (idx 5): sub-headers
  Row 7+ (idx 6+): data

Required columns (0-based index):
  A=0  AI Agent?
  B=1  System / Tool
  C=2  Level (L1-L4)
  D=3  Step Number / Seq
  E=4  Step Name
  F=5  Step Description
  G=6  Step Type (includes decision routing)
  H=7  Automated?
  Q=16 RACI – Responsible (R)
  R=17 RACI – Accountable (A)
  S=18 RACI – Contributing (C)
  T=19 RACI – Informed (I)
"""
import io
import re
import openpyxl

REQUIRED_SHEET = 'Process Design'
HEADER_ROW_IDX = 4   # row 5 in Excel (0-indexed) — main headers
DATA_START_IDX = 6   # row 7 in Excel — skip main header + sub-header row

C_AI, C_SYS, C_LVL, C_SEQ, C_NAME = 0, 1, 2, 3, 4
C_DESC, C_TYPE, C_AUTO = 5, 6, 7
C_R, C_A, C_CONT, C_INF = 16, 17, 18, 19   # Q=Responsible R=Accountable S=Contributing T=Informed

REQUIRED_COLUMNS = {
    C_AI:   ("A",  "AI Agent"),
    C_SYS:  ("B",  "System / Tool"),
    C_LVL:  ("C",  "Level"),
    C_SEQ:  ("D",  "Step Number / Seq"),
    C_NAME: ("E",  "Step Name"),
    C_DESC: ("F",  "Step Description"),
    C_TYPE: ("G",  "Step Type"),
    C_AUTO: ("H",  "Automated"),
    C_R:    ("Q",  "RACI – Responsible"),
    C_A:    ("R",  "RACI – Accountable"),
    C_CONT: ("S",  "RACI – Contributing"),
    C_INF:  ("T",  "RACI – Informed"),
}

L1_COLORS = {
    '1': '#7C3AED',
    '2': '#0E7490',
    '3': '#0891B2',
    '4': '#16A34A',
    '5': '#D97706',
    '6': '#2563EB',
    '7': '#0F766E',
}


def validate_excel(file_content: bytes):
    """Returns (ok: bool, error: str|None, missing_columns: list[str])."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    except Exception as e:
        return False, f"Cannot read Excel file: {e}", []

    if REQUIRED_SHEET not in wb.sheetnames:
        available = ', '.join(wb.sheetnames)
        return False, (
            f"Sheet '{REQUIRED_SHEET}' not found. "
            f"Available sheets: {available}"
        ), []

    ws = wb[REQUIRED_SHEET]
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) <= HEADER_ROW_IDX:
        return False, "File has fewer than 5 rows — main header expected at row 5", []

    header_row = all_rows[HEADER_ROW_IDX]
    missing = []
    for col_idx, (col_letter, col_name) in REQUIRED_COLUMNS.items():
        val = header_row[col_idx] if col_idx < len(header_row) else None
        if not val or str(val).strip() == '':
            missing.append(f"Column {col_letter} – {col_name}")

    if missing:
        return False, "Missing required columns in the header row (row 5)", missing

    return True, None, []


def parse_excel(file_content: bytes) -> list:
    """Parse Excel bytes and return ALL_PROCESSES list."""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb[REQUIRED_SHEET]
    all_rows = list(ws.iter_rows(values_only=True))

    def v(row, col):
        if col >= len(row):
            return ''
        val = row[col]
        s = str(val).strip() if val is not None else ''
        return s.replace('–', '-').replace('—', '-').replace('‘', "'").replace('→', '->')

    def parse_step_type(raw):
        if not raw:
            return 'Process', ''
        lines = [l.strip() for l in raw.split('\n') if l.strip()]
        if not lines:
            return 'Process', ''
        first = lines[0]
        is_dec = ('decision' in first.lower()
                  or any(re.match(r'if\s+(yes|no)', l, re.I) for l in lines))
        if is_dec:
            outcomes = []
            for line in lines[1:]:
                m = re.match(r'If\s+(yes|no)[,.]?\s+(?:proceed to|return to)\s+(.+)', line, re.I)
                if m:
                    target = m.group(2).strip().rstrip('.')
                    target = re.sub(r'^step\s+', '', target, flags=re.I).strip()
                    outcomes.append(m.group(1).capitalize() + ' - proceed to ' + target)
                elif re.match(r'if\s+(yes|no)', line, re.I):
                    cleaned = re.sub(r'^if\s+', '', line, flags=re.I)
                    outcomes.append(cleaned)
            return 'Decision', ' | '.join(outcomes)
        if 'start' in first.lower():
            return 'Start', ''
        return 'Process', ''

    rows_data = []
    for raw in all_rows[DATA_START_IDX:]:
        seq = v(raw, C_SEQ)
        name = v(raw, C_NAME)
        if not seq or not name:
            continue
        if not re.match(r'^\d[\d.]*$', seq):
            continue
        dots = seq.count('.')
        lvl = {0: 'L1', 1: 'L2', 2: 'L3', 3: 'L4'}.get(dots)
        if lvl is None:
            continue

        raw_type = v(raw, C_TYPE)
        step_type, outcomes = parse_step_type(raw_type)
        r_val = v(raw, C_R)
        a_val = v(raw, C_A)
        c_val = v(raw, C_CONT)
        i_val = v(raw, C_INF)

        if 'automated' in r_val.lower() and step_type == 'Process':
            step_type = 'Automated'

        ai_flag = v(raw, C_AI)
        auto_val = v(raw, C_AUTO)
        sys_val = v(raw, C_SYS)
        is_ai = (bool(ai_flag) and ai_flag.lower() not in ('', 'n', 'no', 'none', 'future opportunity')) \
                or 'AI Agent' in r_val \
                or ('partial' in auto_val.lower() and 'agent' in sys_val.lower())

        rows_data.append({
            'lvl': lvl, 'seq': seq, 'name': name,
            'desc': v(raw, C_DESC), 'step_type': step_type, 'outcomes': outcomes,
            'sys': v(raw, C_SYS), 'r': r_val, 'a': a_val, 'c': c_val, 'i': i_val,
            'is_ai': is_ai,
        })

    # Inject synthetic L3 parents for orphaned L4s
    l3_seqs = {r['seq'] for r in rows_data if r['lvl'] == 'L3'}
    orphan_added = set()
    extra = []
    for r in rows_data:
        if r['lvl'] == 'L4':
            parent = '.'.join(r['seq'].split('.')[:3])
            if parent not in l3_seqs and parent not in orphan_added:
                orphan_added.add(parent)
                extra.append({
                    'lvl': 'L3', 'seq': parent, 'name': 'Approval & Sign-off',
                    'desc': '', 'step_type': 'Process', 'outcomes': '',
                    'sys': r['sys'], 'r': r['r'], 'a': r['a'], 'c': r['c'], 'i': r['i'],
                    'is_ai': False,
                })
    rows_data.extend(extra)
    rows_data.sort(
        key=lambda r: [int(x) for x in r['seq'].split('.')] + [0] * (4 - r['seq'].count('.') - 1)
    )

    def make_node(r, children=None):
        node = {
            'seq': r['seq'], 'name': r['name'], 'description': r['desc'],
            'step_type': r['step_type'], 'system_tool': r['sys'],
            'raci': {'r': r['r'], 'a': r['a'], 'c': r['c'], 'i': r['i']},
            'decision_outcomes': r['outcomes'] if r['outcomes'] else None,
            'is_ai': r.get('is_ai', False),
        }
        if children is not None:
            node['children'] = children
        return node

    l1_map, l2_map, l3_map = {}, {}, {}

    for r in rows_data:
        seq = r['seq']
        if r['lvl'] == 'L1':
            raw_id = re.sub(r'[^a-z0-9]+', '-', r['name'].lower()).strip('-')
            l1_map[seq] = {
                'id': raw_id, 'l1_seq': seq, 'l1_name': r['name'],
                'l1_description': r['desc'],
                'l1_color': L1_COLORS.get(seq, '#475569'),
                'raci': {'r': r['r'], 'a': r['a']},
                'system_tool': r['sys'], 'stages': [],
            }
        elif r['lvl'] == 'L2':
            l1_key = seq.split('.')[0]
            rec = {
                'id': seq, 'seq': seq, 'name': r['name'], 'description': r['desc'],
                'step_type': r['step_type'], 'system_tool': r['sys'],
                'raci': {'r': r['r'], 'a': r['a'], 'c': r['c'], 'i': r['i']},
                'steps': [],
            }
            l2_map[seq] = rec
            if l1_key in l1_map:
                l1_map[l1_key]['stages'].append(rec)
        elif r['lvl'] == 'L3':
            l2_key = '.'.join(seq.split('.')[:2])
            rec = make_node(r, children=[])
            l3_map[seq] = rec
            if l2_key in l2_map:
                l2_map[l2_key]['steps'].append(rec)
        elif r['lvl'] == 'L4':
            l3_key = '.'.join(seq.split('.')[:3])
            if l3_key in l3_map:
                l3_map[l3_key]['children'].append(make_node(r))

    return list(l1_map.values())


COLUMNS_META = [
    {"col": "A",  "index": C_AI,   "name": "AI Agent",                      "description": "Flag for AI Agent steps (e.g. X / Y / Yes)"},
    {"col": "B",  "index": C_SYS,  "name": "System / Tool",                 "description": "System or tool used in the step"},
    {"col": "C",  "index": C_LVL,  "name": "Level",                         "description": "Process level (L1, L2, L3, L4)"},
    {"col": "D",  "index": C_SEQ,  "name": "Step Number / Seq",              "description": "Sequence number (e.g. 1, 1.1, 1.1.1, 1.1.1.1)"},
    {"col": "E",  "index": C_NAME, "name": "Step Name",                      "description": "Name of the process step"},
    {"col": "F",  "index": C_DESC, "name": "Step Description",               "description": "Detailed description of the step"},
    {"col": "G",  "index": C_TYPE, "name": "Step Type",                      "description": "Process, Decision, Start, or Automated — Decisions include If yes/no routing"},
    {"col": "H",  "index": C_AUTO, "name": "Automated",                      "description": "Whether the step is fully automated (Y/N/Partial)"},
    {"col": "Q",  "index": C_R,    "name": "RACI – Responsible (R)",    "description": "Who is responsible for executing the step"},
    {"col": "R",  "index": C_A,    "name": "RACI – Accountable (A)",    "description": "Who is ultimately accountable"},
    {"col": "S",  "index": C_CONT, "name": "RACI – Contributing (C)", "description": "Who contributes to the step"},
    {"col": "T",  "index": C_INF,  "name": "RACI – Informed (I)",     "description": "Who is informed of the outcome"},
]
